"""Rebuilds "Game Submissions.xlsx" with the header row/columns the site
expects. Re-run this any time the column list needs to change - it does not
touch any data rows already in the file.
"""

import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "Game Submissions.xlsx")

COLUMNS = [
    "Timestamp",
    "Approval Status",
    "Mod",
    "Date Played",
    "Map",
    "Match Length",
    "Team 1 Faction",
    "Team 1 Commander",
    "Team 1 Thug 1",
    "Team 1 Thug 2",
    "Team 1 Thug 3",
    "Team 1 Thug 4",
    "Team 2 Faction",
    "Team 2 Commander",
    "Team 2 Thug 1",
    "Team 2 Thug 2",
    "Team 2 Thug 3",
    "Team 2 Thug 4",
    "Winner",
    "Verification Type",
    "YouTube Link",
    "Screenshot Link",
    "Submitted By",
    "Reviewer Notes",
]

APPROVAL_STATUSES = ["Pending", "Approved", "Rejected"]
DATA_VALIDATION_ROWS = 1000


def build():
    if os.path.exists(OUTPUT_PATH):
        wb = load_workbook(OUTPUT_PATH)
        ws = wb.active
        ws.delete_rows(1, ws.max_row)
    else:
        wb = Workbook()
        ws = wb.active

    ws.title = "Submissions"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="B5179E", end_color="B5179E", fill_type="solid")

    for col_index, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_index, value=name)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = max(14, len(name) + 2)

    ws.freeze_panes = "A2"

    approval_col_letter = ws.cell(row=1, column=COLUMNS.index("Approval Status") + 1).column_letter
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(APPROVAL_STATUSES)}"',
        allow_blank=True,
    )
    dv.add(f"{approval_col_letter}2:{approval_col_letter}{DATA_VALIDATION_ROWS}")
    ws.add_data_validation(dv)

    wb.save(OUTPUT_PATH)
    print(f"Wrote {len(COLUMNS)} columns to {OUTPUT_PATH}")


if __name__ == "__main__":
    build()
